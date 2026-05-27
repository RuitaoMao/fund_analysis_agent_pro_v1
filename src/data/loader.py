"""Excel 读取层。

这个文件只负责读原始 Excel，不负责具体分析。
"""

from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

from src.data.cleaner import clean_size_df, clean_holding_df, clean_performance_df


def read_excel_fast(path: Path) -> pd.DataFrame:
    """用 openpyxl read_only 模式快速读取 Excel。

    pandas.read_excel 在较大的 xlsx 上有时比较慢；
    这里用 openpyxl 流式读取，再转成 DataFrame，适合当前教学数据。
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    data = list(rows)
    return pd.DataFrame(data, columns=header)


class ExcelDataLoader:
    """读取三张 Excel，并调用 cleaner 统一字段。"""

    def __init__(self, raw_data_dir: Path):
        self.raw_data_dir = raw_data_dir

    def load_clean_data(self) -> dict[str, pd.DataFrame]:
        """读取并清洗规模、持仓、业绩三张表。"""
        size_path = self.raw_data_dir / "规模.xlsx"
        holding_path = self.raw_data_dir / "持仓.xlsx"
        perf_path = self.raw_data_dir / "业绩.xlsx"

        if not size_path.exists() or not holding_path.exists() or not perf_path.exists():
            raise FileNotFoundError(
                "请确认 data/raw/ 下存在 规模.xlsx、持仓.xlsx、业绩.xlsx 三个文件。"
            )

        size_raw = read_excel_fast(size_path)
        holding_raw = read_excel_fast(holding_path)
        perf_raw = read_excel_fast(perf_path)

        return {
            "fund_size": clean_size_df(size_raw),
            "fund_holding": clean_holding_df(holding_raw),
            "fund_performance": clean_performance_df(perf_raw),
        }
